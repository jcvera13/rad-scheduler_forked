"""
exporter.py — Export Layer for Fair Radiology Scheduling

Outputs:
  - Excel (.xlsx): formatted date × shift grid with staff names
  - CSV: flat (date, shift, staff) for programmatic review
  - Fairness audit report (.txt): per-radiologist counts, weighted workload, CV
    including per-shift breakdown (M0 CV, M1 CV, etc.)

Usage:
  from src.exporter import export_to_csv, export_to_excel, export_fairness_report
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

Schedule = Dict[str, List[Tuple[str, str]]]   # date_str → [(shift, name)]


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------

def export_to_csv(
    schedule: Schedule,
    output_path: Path,
    include_shift: bool = True,
) -> None:
    """
    Export schedule to flat CSV: date, shift, staff.

    Args:
        schedule:      {date_str: [(shift_name, person_name), ...]}
        output_path:   .csv file path
        include_shift: If True, include shift column
    """
    import csv
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="") as f:
        if include_shift:
            writer = csv.DictWriter(f, fieldnames=["date", "shift", "staff"])
            writer.writeheader()
            for date_str, assignments in sorted(schedule.items()):
                for shift_name, person_name in assignments:
                    writer.writerow({"date": date_str, "shift": shift_name, "staff": person_name})
        else:
            writer = csv.DictWriter(f, fieldnames=["date", "staff"])
            writer.writeheader()
            for date_str, assignments in sorted(schedule.items()):
                for _, person_name in assignments:
                    writer.writerow({"date": date_str, "staff": person_name})

    logger.info(f"CSV exported → {output_path}")


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_to_excel(
    schedule: Schedule,
    output_path: Path,
    pivot: bool = True,
    shift_order: Optional[List[str]] = None,
    name_to_initials: Optional[Dict[str, str]] = None,
) -> None:
    """
    Export schedule to formatted Excel grid.

    Pivot mode (default): rows=date, columns=shift, cells=staff name (or initials).
    Flat mode: rows=(date, shift, staff).

    Args:
        schedule:     {date_str: [(shift_name, person_name), ...]}
        output_path:  .xlsx file path
        pivot:        If True, create date × shift grid
        shift_order:  Column order for pivot (defaults to appearance order)
        name_to_initials:  If provided, cells show initials (e.g. DA, JCV) instead of full names
    """
    import pandas as pd

    output_path.parent.mkdir(parents=True, exist_ok=True)
    name_map = name_to_initials or {}

    rows = []
    for date_str, assignments in sorted(schedule.items()):
        for shift_name, person_name in assignments:
            display_name = name_map.get(person_name, person_name)
            rows.append({"Date": date_str, "Shift": shift_name, "Staff": display_name})

    df = pd.DataFrame(rows)
    if df.empty:
        df.to_excel(output_path, index=False)
        return

    if pivot:
        # Pivot: Date × Shift → Staff name
        grid = df.pivot_table(
            index="Date",
            columns="Shift",
            values="Staff",
            aggfunc=lambda x: "; ".join(x),  # handles multi-assign per cell
        )
        # Reorder columns if provided
        if shift_order:
            available = [s for s in shift_order if s in grid.columns]
            rest = [s for s in grid.columns if s not in shift_order]
            grid = grid[available + rest]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            grid.to_excel(writer, sheet_name="Schedule")
            _format_excel_grid(writer, "Schedule", grid)
    else:
        df.to_excel(output_path, index=False)

    logger.info(f"Excel exported → {output_path}")


def _format_excel_grid(writer: Any, sheet_name: str, grid: Any) -> None:
    """Apply basic formatting to Excel grid: column widths, header bold."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        # Alternate row shading
        from openpyxl.styles import PatternFill as PF
        alt = PF("solid", fgColor="EBF3FB")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = alt

    except Exception as e:
        logger.warning(f"Excel formatting failed (non-critical): {e}")


# ---------------------------------------------------------------------------
# Fairness Report
# ---------------------------------------------------------------------------

def export_fairness_report(
    metrics: Dict[str, Any],
    output_path: Path,
    top_n: int = 3,
    bottom_n: int = 3,
    pool_label: str = "",
    target_cv: float = 10.0,
) -> None:
    """
    Export full fairness audit report (text format).

    Includes:
      - Overall CV, mean, std, min, max
      - Per-radiologist weighted counts with deviation from mean
      - Per-shift CV breakdown (M0 CV, M1 CV, etc.)
      - Top/bottom N assigned
      - UNFILLED slot count
      - Pass/fail vs target CV

    Args:
        metrics:     Output of engine.calculate_fairness_metrics()
        output_path: .txt file path
        top_n:       Number of most-assigned to highlight
        bottom_n:    Number of least-assigned to highlight
        pool_label:  Label for the rotation pool (e.g. 'Mercy Weekday')
        target_cv:   CV target (default 10.0%)
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wc = metrics.get("weighted_counts", metrics.get("counts", {}))
    rc = metrics.get("counts", {})
    hc = metrics.get("hours_counts", {})
    per_shift = metrics.get("per_shift", {})
    per_shift_cv = metrics.get("per_shift_cv", {})
    mean_val = metrics.get("mean", 0)
    cv = metrics.get("cv", 0)
    hours_mean = metrics.get("hours_mean", 0)
    hours_cv = metrics.get("hours_cv", 0)
    unfilled = metrics.get("unfilled", 0)

    sorted_names = sorted(wc.keys(), key=lambda n: wc.get(n, 0), reverse=True)
    pass_fail = "✓ PASS" if cv < target_cv else "✗ FAIL"
    hours_pass_fail = "✓ PASS" if hours_cv < target_cv else "✗ FAIL"

    sep = "=" * 70

    lines = [
        sep,
        f"  FAIRNESS AUDIT REPORT{(' — ' + pool_label) if pool_label else ''}",
        sep,
        "",
        f"  Overall Weighted CV:   {cv:.2f}%  (target <{target_cv:.0f}%)  {pass_fail}",
        f"  Mean weighted load:    {mean_val:.2f}",
        f"  Std Dev:               {metrics.get('std', 0):.2f}",
        f"  Min / Max:             {metrics.get('min', 0):.2f} / {metrics.get('max', 0):.2f}",
        "",
        f"  Hours Assigned CV:     {hours_cv:.2f}%  (target <{target_cv:.0f}%)  {hours_pass_fail}",
        f"  Mean hours assigned:   {hours_mean:.2f}",
        f"  Hours Std Dev:         {metrics.get('hours_std', 0):.2f}",
        f"  Unfilled slots:        {unfilled}",
        "",
        "─" * 70,
        "  Per-Radiologist Weighted Workload & Hours",
        "─" * 70,
        f"  {'Name':<24} {'Wt Load':>8} {'Hours':>8} {'Raw':>6} {'% Mean':>8}  {'Δ Mean':>8}",
    ]

    for name in sorted_names:
        wt = wc.get(name, 0)
        raw = rc.get(name, 0)
        hrs = hc.get(name, 0)
        pct = (wt / mean_val * 100) if mean_val else 0
        delta = wt - mean_val
        flag = "  ← ↑ over" if delta > mean_val * 0.20 else ("  ← ↓ under" if delta < -mean_val * 0.20 else "")
        lines.append(
            f"  {name:<24} {wt:>8.2f} {hrs:>8.1f} {raw:>6d} {pct:>7.1f}% {delta:>+9.2f}{flag}"
        )

    lines += [
        "",
        "─" * 70,
        "  Per-Shift CV Breakdown",
        "─" * 70,
    ]

    if per_shift_cv:
        for shift_name, cv_val in sorted(per_shift_cv.items()):
            flag = " ✗" if cv_val >= target_cv else " ✓"
            lines.append(f"  {shift_name:<16} CV={cv_val:6.2f}%{flag}")
    else:
        lines.append("  (no per-shift breakdown available)")

    lines += [
        "",
        "─" * 70,
        "  Top Assigned",
        "─" * 70,
    ]
    for name in sorted_names[:top_n]:
        wt = wc.get(name, 0)
        pct = (wt / mean_val * 100) if mean_val else 0
        lines.append(f"  {name:<24} {wt:.2f} ({pct:.1f}% of mean)")

    lines += [
        "",
        "─" * 70,
        "  Bottom Assigned",
        "─" * 70,
    ]
    for name in sorted_names[-bottom_n:][::-1]:
        wt = wc.get(name, 0)
        pct = (wt / mean_val * 100) if mean_val else 0
        lines.append(f"  {name:<24} {wt:.2f} ({pct:.1f}% of mean)")

    lines.append("")
    lines.append(sep)

    report_text = "\n".join(lines)
    with open(output_path, "w") as f:
        f.write(report_text)

    logger.info(f"Fairness report exported → {output_path}")
    return report_text
