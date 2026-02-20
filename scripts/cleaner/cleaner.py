"""
clean_ha_values.py
------------------
Interactively assigns HA (actual hours worked) values to tasks in a
Qgenda-style scheduling spreadsheet where HA = 0 or is blank.

Usage:
    python clean_ha_values.py <input_file.xlsx>

Outputs:
    - <filename>_cleaned.xlsx      — updated workbook (formatting preserved)
    - schedule_ha_config.json      — shared task→HA mapping log (same dir as script)
    - Summary report printed to console
"""

import sys
import os
import json
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Config ──────────────────────────────────────────────────────────────────

HEADER_ANCHOR   = "Date"        # Value in col A of the true header row
TASK_COL_IDX    = 3             # 0-based: column D = "Task"
HA_COL_IDX      = 8             # 0-based: column I = "HA"
EMPTY_HA_VALS   = {0, None, ""} # Treat these as "unassigned"
MIN_HOURS, MAX_HOURS = 0, 12
CONFIG_FILENAME = "schedule_ha_config.json"


# ── Helpers ─────────────────────────────────────────────────────────────────

def find_header_row(ws):
    """Return 1-based row number of the true column header row."""
    for row in ws.iter_rows():
        if row[0].value == HEADER_ANCHOR:
            return row[0].row
    raise ValueError(f"Could not find header row (looking for '{HEADER_ANCHOR}' in col A).")


def load_json_config(config_path):
    """Load config if it exists and has valid structure, else return blank shell."""
    if os.path.exists(config_path):
        try:
            with open(config_path, "r") as f:
                data = json.load(f)
            if isinstance(data.get("task_ha_map"), dict):
                return data
            else:
                print("  ⚠  Config found but has unexpected format — starting fresh.")
        except (json.JSONDecodeError, IOError) as e:
            print(f"  ⚠  Could not read config ({e}) — starting fresh.")
    return {"task_ha_map": {}, "history": []}


def save_json_config(config, config_path):
    with open(config_path, "w") as f:
        json.dump(config, f, indent=2)


def prompt_ha_value(task_name, count, example_dates, existing_val=None):
    """
    Prompt user for an HA value for a task group.
    Returns float|int value, or None if skipped.
    """
    print("\n" + "─" * 65)
    print(f"  TASK  : {task_name}")
    print(f"  ROWS  : {count} occurrence(s) with HA = 0 / blank")
    print(f"  DATES : {', '.join(example_dates[:5])}" +
          (" …" if len(example_dates) > 5 else ""))
    if existing_val is not None:
        print(f"  CONFIG: Previously assigned HA = {existing_val}  (press Enter to reuse)")
    print()

    while True:
        hint = f"[Enter={existing_val}] " if existing_val is not None else ""
        raw = input(f"  Enter HA hours ({MIN_HOURS}–{MAX_HOURS}), or 's' to skip: {hint}").strip()

        if raw.lower() == "s":
            return None

        # Reuse existing config value on bare Enter
        if raw == "" and existing_val is not None:
            return existing_val

        try:
            val = float(raw)
            val = val if val != int(val) else int(val)  # keep 1.5, simplify 9.0→9
            if MIN_HOURS <= val <= MAX_HOURS:
                return val
            else:
                print(f"  ✗  Value must be between {MIN_HOURS} and {MAX_HOURS}.")
        except ValueError:
            print("  ✗  Please enter a number or 's' to skip.")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    # ── 1. Argument & file handling ─────────────────────────────────────────
    if len(sys.argv) < 2:
        print("Usage: python clean_ha_values.py <input_file.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        sys.exit(1)

    script_dir  = os.path.dirname(os.path.abspath(__file__))
    base_name   = os.path.splitext(os.path.basename(input_path))[0]
    ext         = os.path.splitext(input_path)[1]
    output_path = os.path.join(script_dir, base_name + "_cleaned" + ext)
    config_path = os.path.join(script_dir, CONFIG_FILENAME)

    print(f"\n{'═'*65}")
    print(f"  HA Column Cleaner")
    print(f"{'═'*65}")
    print(f"  Input  : {input_path}")
    print(f"  Output : {output_path}")
    print(f"  Config : {config_path}")
    print(f"{'═'*65}\n")

    # ── 2. Load workbook ─────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # ── 3. Detect & prompt for existing config ───────────────────────────────
    config       = load_json_config(config_path)
    existing_map = config.get("task_ha_map", {})

    if existing_map:
        print(f"  ✓ Existing config detected: {CONFIG_FILENAME}")
        print(f"    Contains {len(existing_map)} task→HA mapping(s).")
        history = config.get("history", [])
        if history:
            last = history[-1]
            print(f"    Last updated : {last.get('run_timestamp', 'unknown')}")
            print(f"    Source file  : {last.get('input_file', 'unknown')}")
        print()

        while True:
            choice = input("  Load existing mappings as defaults? (y/n): ").strip().lower()
            if choice in ("y", "n"):
                break
            print("  Please enter 'y' or 'n'.")

        if choice == "n":
            print("  → Ignoring existing mappings. All tasks will be prompted fresh.\n")
            task_ha_map = {}
        else:
            print("  → Mappings loaded. Press Enter at any prompt to reuse a value.\n")
            task_ha_map = dict(existing_map)
    else:
        print(f"  No existing config found — starting fresh.\n")
        task_ha_map = {}

    # ── 4. Scan: collect all tasks + flag HA=0/blank rows ───────────────────
    header_row = find_header_row(ws)
    all_tasks  = defaultdict(lambda: {"total": 0, "zero_rows": [], "ha_vals": set()})

    for row in ws.iter_rows(min_row=header_row + 1):
        date_val = row[0].value
        task_val = row[TASK_COL_IDX].value
        ha_cell  = row[HA_COL_IDX]
        ha_val   = ha_cell.value

        if task_val is None or date_val is None:
            continue
        if str(date_val).strip().lower() in ("totals", ""):
            continue

        task_name = str(task_val).strip()
        all_tasks[task_name]["total"] += 1
        all_tasks[task_name]["ha_vals"].add(ha_val)

        if ha_val in EMPTY_HA_VALS:
            all_tasks[task_name]["zero_rows"].append((ha_cell, str(date_val)))

    needs_assignment = {t: d for t, d in all_tasks.items() if d["zero_rows"]}
    already_assigned = {t: d for t, d in all_tasks.items() if not d["zero_rows"]}

    print(f"  Scan complete: {len(all_tasks)} unique task types across {ws.max_row - header_row} data rows.")
    print(f"  → {len(already_assigned)} tasks already have non-zero HA values.")
    print(f"  → {len(needs_assignment)} tasks have HA = 0 / blank rows to review.\n")

    if not needs_assignment:
        print("  ✓ Nothing to do — all HA values are already assigned.")
        sys.exit(0)

    # ── 5. Interactive assignment loop ───────────────────────────────────────
    session_assignments = {}
    skipped_tasks       = []

    sorted_tasks = sorted(needs_assignment.items(), key=lambda x: x[0].lower())
    total_tasks  = len(sorted_tasks)

    print(f"  You will be prompted for each of the {total_tasks} task group(s).")
    print("  Enter a number 0–12, or 's' to skip a task.\n")

    for idx, (task_name, data) in enumerate(sorted_tasks, 1):
        print(f"  [{idx}/{total_tasks}]", end="")
        example_dates = [d for _, d in data["zero_rows"]]
        existing_val  = task_ha_map.get(task_name)

        assigned = prompt_ha_value(
            task_name,
            count=len(data["zero_rows"]),
            example_dates=example_dates,
            existing_val=existing_val,
        )

        if assigned is None:
            skipped_tasks.append(task_name)
            print(f"    → Skipped.")
        else:
            session_assignments[task_name] = assigned
            task_ha_map[task_name] = assigned
            for (ha_cell, _) in data["zero_rows"]:
                ha_cell.value = assigned
            changed = " (updated)" if existing_val is not None and existing_val != assigned else ""
            print(f"    → Assigned HA = {assigned} to {len(data['zero_rows'])} row(s).{changed}")

    # ── 6. Save cleaned workbook ─────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n  ✓ Saved cleaned workbook → {output_path}")

    # ── 7. Update JSON config ────────────────────────────────────────────────
    config["task_ha_map"] = task_ha_map
    config.setdefault("history", []).append({
        "run_timestamp" : datetime.now().isoformat(timespec="seconds"),
        "input_file"    : os.path.basename(input_path),
        "output_file"   : os.path.basename(output_path),
        "assigned"      : session_assignments,
        "skipped"       : skipped_tasks,
    })
    save_json_config(config, config_path)
    print(f"  ✓ Updated config         → {config_path}")

    # ── 8. Summary report ────────────────────────────────────────────────────
    print(f"\n{'═'*65}")
    print("  SUMMARY REPORT")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'═'*65}")

    # Section A: Already-assigned tasks (non-zero before this run)
    print(f"\n  [A] TASKS WITH EXISTING HA VALUES  ({len(already_assigned)} tasks)\n")
    print(f"  {'Task':<55} {'HA Value(s)'}")
    print(f"  {'-'*54} {'-'*10}")
    for task_name in sorted(already_assigned.keys(), key=str.lower):
        d        = already_assigned[task_name]
        vals_str = ", ".join(str(v) for v in sorted(d["ha_vals"], key=lambda x: (x is None, x or 0)))
        print(f"  {task_name:<55} {vals_str}  ({d['total']} rows)")

    # Section B: Tasks assigned this session
    if session_assignments:
        print(f"\n  [B] TASKS ASSIGNED THIS SESSION  ({len(session_assignments)} tasks)\n")
        print(f"  {'Task':<55} {'HA Assigned':<12} {'Rows Updated'}")
        print(f"  {'-'*54} {'-'*11} {'-'*12}")
        for task_name, ha_val in sorted(session_assignments.items(), key=lambda x: x[0].lower()):
            n_rows = len(needs_assignment[task_name]["zero_rows"])
            print(f"  {task_name:<55} {str(ha_val):<12} {n_rows}")

    # Section C: Skipped tasks
    if skipped_tasks:
        print(f"\n  [C] SKIPPED TASKS  ({len(skipped_tasks)} tasks — HA values unchanged)\n")
        print(f"  {'Task':<55} {'Rows Still = 0'}")
        print(f"  {'-'*54} {'-'*14}")
        for task_name in sorted(skipped_tasks, key=str.lower):
            n_rows = len(needs_assignment[task_name]["zero_rows"])
            print(f"  {task_name:<55} {n_rows}")

    # Totals
    total_updated = sum(len(needs_assignment[t]["zero_rows"]) for t in session_assignments)
    total_skipped = sum(len(needs_assignment[t]["zero_rows"]) for t in skipped_tasks)
    print(f"\n{'─'*65}")
    print(f"  Total rows updated : {total_updated}")
    print(f"  Total rows skipped : {total_skipped}")
    print(f"  Output file        : {output_path}")
    print(f"  Config file        : {config_path}")
    print(f"{'═'*65}\n")


if __name__ == "__main__":
    main()
